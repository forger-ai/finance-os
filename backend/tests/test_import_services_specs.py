from __future__ import annotations

import csv
import io
from datetime import date, datetime, time, timezone

import pytest
from openpyxl import Workbook
from pypdf import PdfWriter
from sqlmodel import Session, select

from app.database import engine
from app.models import CategoryKind, Movement, MovementSource, Setting
from app.services import document_preprocessor
from app.services import import_movements as import_service
from app.services import settings as settings_service
from app.services import xlsx_to_csv as xlsx_service
from app.services.classification import ClassificationError, resolve_movement_classification
from app.services.classification_memory import build_classification_memory, memory_index
from app.services.document_preprocessor import preprocess_document
from app.services.import_movements import (
    has_recognizable_schema,
    import_movements_from_csv,
    import_movements_structured,
    parse_csv_text,
)
from app.services.integrity import find_classification_mismatches
from app.services.settings import (
    PRIMARY_CURRENCY_CODE_KEY,
    currency_format_by_code,
    currency_formats,
    get_primary_currency_code,
    settings_payload,
    update_primary_currency_code,
    validate_currency_code,
)
from app.services.xlsx_to_csv import xlsx_to_csv
from app.utils import (
    UTC,
    isoformat_z,
    normalize_key,
    normalize_text,
    parse_action_date,
    parse_boolean,
    parse_date_input,
    to_cents,
    to_pesos,
    to_positive_cents,
)
from tests.helpers import seed_category, seed_subcategory

pytestmark = pytest.mark.bdd


def test_csv_import_aliases_empty_input_row_failures_memory_and_duplicates(
    session: Session,
) -> None:
    # Given a category tree and reviewed history that the importer can reuse.
    expenses = seed_category(session, name="Gastos")
    groceries = seed_subcategory(session, name="Supermercado", category=expenses)
    transport = seed_subcategory(session, name="Transporte", category=expenses)
    _seed_movement(
        session,
        category_id=expenses.id,
        subcategory_id=transport.id,
        business="Metro Recarga",
        reviewed=True,
        source_row="history",
    )

    # Empty files are accepted as no-op imports and are not recognizable schemas.
    assert parse_csv_text("") == []
    assert has_recognizable_schema("") is False
    empty = import_movements_from_csv(session, "", file_label="empty.csv")
    assert empty.to_dict() == {
        "file": "empty.csv",
        "inserted": 0,
        "duplicate": 0,
        "failed": 0,
        "errors": [],
        "rows": [],
    }

    # When a bank CSV uses Spanish aliases and has mixed valid/invalid rows.
    csv_text = _csv_text(
        [
            [
                "Fecha Movimiento",
                "Monto Cargo ($)",
                "Monto Abono",
                "Descripción",
                "Subcategoría",
                "Origen",
                "Revisado",
                "Fecha Contable",
            ],
            [
                "01/05/26",
                "1.500,50",
                "",
                "Jumbo compra",
                "Supermercado",
                "BANK",
                "yes",
                "02-05-2026",
            ],
            ["2026-05-03", "", "2000", "Metro Recarga", "", "MANUAL", "no", ""],
            ["not-a-date", "100", "", "Fecha mala", "", "MANUAL", "false", ""],
            ["2026-05-04", "abc", "", "Monto malo", "", "MANUAL", "false", ""],
            ["2026-05-05", "100", "", "Bool malo", "", "MANUAL", "maybe", ""],
            ["2026-05-06", "100", "", "Fuente mala", "", "WIRE", "false", ""],
            ["2026-05-07", "100", "", "Sub mala", "Fantasma", "MANUAL", "false", ""],
        ]
    )

    # Then aliases, charge/credit amounts, memory classification, and failures are reported.
    assert has_recognizable_schema(csv_text) is True
    outcome = import_movements_from_csv(session, csv_text, file_label="bank.csv")
    assert outcome.inserted == 2
    assert outcome.failed == 5
    status_by_row = {row.row: row.status for row in outcome.rows}
    assert status_by_row == {
        2: "inserted",
        3: "inserted",
        4: "failed",
        5: "failed",
        6: "failed",
        7: "failed",
        8: "failed",
    }
    assert {error.row for error in outcome.errors} == {4, 5, 6, 7, 8}
    assert any("Invalid date value" in error.error for error in outcome.errors)
    assert any("Falta el monto" in error.error for error in outcome.errors)
    assert any("Invalid boolean value" in error.error for error in outcome.errors)
    assert any("Origen desconocido" in error.error for error in outcome.errors)
    assert any("Subcategoría desconocida" in error.error for error in outcome.errors)

    inserted = sorted(
        session.exec(select(Movement).where(Movement.source_file == "bank.csv")).all(),
        key=lambda movement: movement.source_row or "",
    )
    assert [(movement.business, movement.amount_cents) for movement in inserted] == [
        ("Jumbo compra", 150050),
        ("Metro Recarga", 200000),
    ]
    assert inserted[0].subcategory_id == groceries.id
    assert inserted[0].reviewed is True
    assert inserted[1].subcategory_id == transport.id

    # And importing the same valid source rows again reports hash duplicates, not failures.
    duplicate = import_movements_from_csv(session, csv_text, file_label="bank.csv")
    assert duplicate.inserted == 0
    assert duplicate.duplicate == 2
    assert duplicate.failed == 5


def test_imports_require_fallback_category_before_csv_or_structured_import(
    session: Session,
) -> None:
    # Given an empty database with no category to fall back to.
    csv_text = "date,amount,reason\n2026-05-01,10,No category\n"

    # Then CSV and structured imports surface the missing setup clearly.
    with pytest.raises(ValueError, match="At least one category"):
        import_movements_from_csv(session, csv_text, file_label="missing-category.csv")

    outcome = import_movements_structured(
        session,
        [{"date": "2026-05-01", "amount": 10, "business": "No category"}],
        file_label="structured.csv",
    )
    assert outcome.inserted == 0
    assert outcome.failed == 1
    assert "At least one category" in outcome.errors[0].error


def test_structured_import_resolves_classification_forms_and_row_failures(
    session: Session,
) -> None:
    # Given multiple category trees, including duplicate subcategory names.
    expenses = seed_category(session, name="Gastos")
    income = seed_category(session, name="Ingresos", kind=CategoryKind.INCOME)
    other_expenses = seed_category(session, name="Otros gastos")
    groceries = seed_subcategory(session, name="Supermercado", category=expenses)
    transport = seed_subcategory(session, name="Transporte", category=expenses)
    salary = seed_subcategory(session, name="Sueldo", category=income)
    seed_subcategory(session, name="Comida", category=expenses)
    seed_subcategory(session, name="Comida", category=other_expenses)

    # When the assistant sends valid id/name/subcategory forms plus invalid rows.
    outcome = import_movements_structured(
        session,
        [
            {
                "_row_number": 10,
                "date": "2026-05-01",
                "amount": "10",
                "business": "By category id",
                "category_id": expenses.id,
                "subcategory": "Supermercado",
            },
            {
                "_row_number": 11,
                "date": "2026-05-02",
                "amount": 20,
                "business": "By category name",
                "category": "ingresos",
                "reason": "Salary",
            },
            {
                "_row_number": 12,
                "date": "2026-05-03",
                "amount": 30,
                "business": "By subcategory id",
                "subcategory_id": salary.id,
            },
            {
                "_row_number": 13,
                "date": "2026-05-04",
                "amount": 40,
                "business": "By unique subcategory",
                "subcategory": "Transporte",
            },
            {
                "_row_number": 14,
                "date": "2026-05-05",
                "amount": 50,
                "business": "Ambiguous subcategory",
                "subcategory": "Comida",
            },
            {
                "_row_number": 15,
                "date": "2026-05-06",
                "amount": 60,
                "business": "Unknown subcategory id",
                "subcategory_id": "missing-subcategory",
            },
            {"_row_number": 16, "date": "", "amount": 70, "business": "Missing date"},
            {"_row_number": 17, "date": "2026-05-08", "business": "Missing amount"},
            {
                "_row_number": 18,
                "date": "2026-05-09",
                "amount": 90,
                "business": "Bad source",
                "source": "WIRE",
            },
        ],
        file_label="assistant.json",
    )

    # Then each row is independent and classification is resolved deterministically.
    assert outcome.inserted == 4
    assert outcome.failed == 5
    assert [row.row for row in outcome.rows if row.status == "inserted"] == [10, 11, 12, 13]
    assert {row.row for row in outcome.rows if row.status == "failed"} == {
        14,
        15,
        16,
        17,
        18,
    }
    assert any("Ambiguous subcategory" in error.error for error in outcome.errors)
    assert any("Unknown subcategory_id" in error.error for error in outcome.errors)
    assert any("Missing date" in error.error for error in outcome.errors)
    assert any("Missing amount" in error.error for error in outcome.errors)
    assert any("Unknown source" in error.error for error in outcome.errors)

    imported = {
        movement.business: movement
        for movement in session.exec(
            select(Movement).where(Movement.source_file == "assistant.json")
        ).all()
    }
    assert imported["By category id"].subcategory_id == groceries.id
    assert imported["By category name"].category_id == income.id
    assert imported["By subcategory id"].category_id == income.id
    assert imported["By unique subcategory"].subcategory_id == transport.id


def test_structured_import_tracks_exact_duplicates_and_similarity_warnings(
    session: Session,
) -> None:
    # Given an existing movement without an import hash and one category.
    expenses = seed_category(session, name="Gastos")
    _seed_movement(
        session,
        category_id=expenses.id,
        business="Cafe Central",
        reason="Latte",
        raw_description="Latte",
        amount_cents=450000,
        source_file=None,
        source_row=None,
        import_hash=None,
    )

    # When a matching semantic row arrives with a stable source fingerprint.
    similar = import_movements_structured(
        session,
        [
            {
                "date": "2026-05-01",
                "amount": 4500,
                "business": "Cafe Central",
                "reason": "Latte",
                "raw_description": "Latte",
                "external_id": "cafe-1",
            }
        ],
        file_label="card.csv",
    )

    # Then it imports once but is marked as a possible duplicate by similarity.
    assert similar.inserted == 1
    inserted_id = similar.rows[0].movement_id
    assert inserted_id is not None
    inserted = session.get(Movement, inserted_id)
    assert inserted is not None
    assert inserted.duplicate_warning == "possible_duplicate"

    # And the same source fingerprint is skipped on the next import.
    duplicate = import_movements_structured(
        session,
        [
            {
                "date": "2026-05-02",
                "amount": 9999,
                "business": "Changed text",
                "external_id": "cafe-1",
            }
        ],
        file_label="card.csv",
    )
    assert duplicate.inserted == 0
    assert duplicate.duplicate == 1
    assert duplicate.rows[0].duplicate_of == inserted_id


def test_classification_memory_settings_and_integrity_service_branches(
    session: Session,
) -> None:
    # Given reviewed and unreviewed history with a missing legacy subcategory reference.
    expenses = seed_category(session, name="Gastos")
    income = seed_category(session, name="Ingresos", kind=CategoryKind.INCOME)
    groceries = seed_subcategory(session, name="Supermercado", category=expenses)
    _seed_movement(
        session,
        category_id=expenses.id,
        subcategory_id=groceries.id,
        business="Jumbo",
        reviewed=True,
        source_row="1",
    )
    _seed_movement(
        session,
        category_id=expenses.id,
        subcategory_id=groceries.id,
        business="Jumbo",
        reviewed=True,
        source_row="2",
    )
    legacy = _seed_movement(
        session,
        category_id=income.id,
        business="Legacy employer",
        reviewed=True,
        source_row="3",
    )
    _force_subcategory_id(session, legacy.id, "deleted-subcategory")
    _seed_movement(
        session,
        category_id=expenses.id,
        business="Ignored pending",
        reviewed=False,
        source_row="4",
    )
    _seed_movement(
        session,
        category_id=expenses.id,
        business=" ",
        reviewed=True,
        source_row="5",
    )

    # Then memory keeps confident reviewed classifications and tolerates stale subcategories.
    memory = build_classification_memory(session, min_confidence=2)
    assert [(entry.business, entry.count) for entry in memory] == [("Jumbo", 2)]
    assert memory_index(memory) == {"jumbo": memory[0]}

    limited = build_classification_memory(session, limit=1)
    assert len(limited) == 1
    legacy = next(
        entry
        for entry in build_classification_memory(session)
        if entry.business == "Legacy employer"
    )
    assert legacy.category_id == income.id
    assert legacy.subcategory_id is None
    _force_category_id(session, "deleted-category", legacy.business)
    assert all(
        entry.business != "Legacy employer"
        for entry in build_classification_memory(session)
    )

    # Classification resolution covers missing category, missing subcategory, and mismatch errors.
    assert resolve_movement_classification(
        session,
        category_id=None,
        subcategory_id=groceries.id,
    ).category.id == expenses.id
    with pytest.raises(ClassificationError, match="Categoría"):
        resolve_movement_classification(session, category_id=None, subcategory_id=None)
    with pytest.raises(ClassificationError, match="Subcategoría"):
        resolve_movement_classification(
            session,
            category_id=expenses.id,
            subcategory_id="missing",
        )
    with pytest.raises(ClassificationError, match="Categoría"):
        resolve_movement_classification(session, category_id="missing", subcategory_id=None)
    with pytest.raises(ClassificationError, match="no pertenece"):
        resolve_movement_classification(
            session,
            category_id=income.id,
            subcategory_id=groceries.id,
        )

    # Settings normalize valid currency codes and fall back from invalid persisted values.
    assert currency_formats() == sorted(currency_formats(), key=lambda item: item["code"])
    assert currency_format_by_code(" usd ") is not None
    assert validate_currency_code(" clp ") == "CLP"
    with pytest.raises(ValueError, match="no soportado"):
        validate_currency_code("XXX")
    assert get_primary_currency_code(session) == "CLP"
    assert update_primary_currency_code(session, "usd") == "USD"
    assert settings_payload(session)["primary_currency_code"] == "USD"
    setting = session.get(Setting, PRIMARY_CURRENCY_CODE_KEY)
    assert setting is not None
    setting.value = "XXX"
    session.add(setting)
    session.commit()
    assert get_primary_currency_code(session) == "CLP"

    # Integrity reports both category mismatches and missing subcategory references.
    mismatches = find_classification_mismatches(session)
    stale = next(item for item in mismatches if item.subcategory_id == "deleted-subcategory")
    assert stale.expected_category_id is None
    session.add(
        Movement(
            date=datetime(2026, 5, 1, tzinfo=timezone.utc),
            accounting_date=datetime(2026, 5, 1, tzinfo=timezone.utc),
            amount_cents=1000,
            business="Bad mismatch",
            reason="Bad mismatch",
            source=MovementSource.MANUAL,
            category_id=income.id,
            subcategory_id=groceries.id,
        )
    )
    session.commit()
    assert any(
        item.subcategory_id == groceries.id and item.expected_category_id == expenses.id
        for item in find_classification_mismatches(session)
    )


def test_xlsx_conversion_uses_active_sheet_and_stringifies_supported_cell_types() -> None:
    # Given a workbook whose active sheet has banner and empty rows.
    workbook = Workbook()
    workbook.active.title = "Ignored"
    workbook.active.append(["Fecha", "Monto"])
    workbook.active.append(["1999-01-01", 999])
    active = workbook.create_sheet("Movements")
    workbook.active = 1
    active.append(["Estado de cuenta mayo 2026"])
    active.append([None, None, None, None])
    active.append(["Fecha", "Fecha contable", "Hora", "Monto", "Nulo"])
    active.append(
        [
            datetime(2026, 5, 1, 13, 45),
            datetime(2026, 5, 2).date(),
            time(9, 30),
            1500.25,
            None,
        ]
    )
    active.append([datetime(2026, 5, 3), datetime(2026, 5, 4).date(), time(8), 0.0])

    # When it is converted to CSV.
    converted = xlsx_to_csv(_workbook_bytes(workbook))
    rows = list(csv.reader(io.StringIO(converted)))

    # Then only the active sheet is used, titles/empty rows are skipped, and values are stable.
    assert rows == [
        ["Fecha", "Fecha contable", "Hora", "Monto", "Nulo"],
        ["2026-05-01T13:45:00", "2026-05-02T00:00:00", "09:30:00", "1500.25", ""],
        ["2026-05-03T00:00:00", "2026-05-04T00:00:00", "08:00:00", "0", ""],
    ]


def test_document_preprocessor_handles_supported_files_truncation_and_errors() -> None:
    # CSV content is decoded as UTF-8, counted before truncation, and may be truncated.
    csv_doc = preprocess_document(
        filename="statement.csv",
        content_type="text/csv",
        data=b"date,amount\n2026-05-01,100\n2026-05-02,200\n",
        max_chars=18,
    )
    assert csv_doc.kind == "csv"
    assert csv_doc.row_count == 2
    assert csv_doc.text.endswith("[Truncated to 18 characters.]")

    with pytest.raises(UnicodeDecodeError):
        preprocess_document(
            filename="bad.csv",
            content_type="text/csv",
            data=b"\xff\xfe",
        )

    # XLSX files are normalized into CSV text before row counting and truncation.
    workbook = Workbook()
    workbook.active.append(["Fecha", "Monto"])
    workbook.active.append(["2026-05-01", 100])
    xlsx_doc = preprocess_document(
        filename="statement.xlsx",
        content_type="application/octet-stream",
        data=_workbook_bytes(workbook),
    )
    assert xlsx_doc.kind == "xlsx_as_csv"
    assert xlsx_doc.row_count == 1
    assert "2026-05-01,100" in xlsx_doc.text

    # PDFs with no selectable text still report page count and a warning.
    pdf_doc = preprocess_document(
        filename="blank.pdf",
        content_type="application/pdf",
        data=_blank_pdf_bytes(),
    )
    assert pdf_doc.kind == "pdf_text"
    assert pdf_doc.page_count == 1
    assert pdf_doc.text == ""
    assert pdf_doc.warning is not None

    # Unsupported binary files are returned as unsupported with no text extraction.
    unsupported = preprocess_document(
        filename="statement.bin",
        content_type="application/octet-stream",
        data=b"\x00\x01",
    )
    assert unsupported.kind == "unsupported_binary"
    assert unsupported.text == ""
    assert unsupported.warning is not None


def test_utility_parsers_cover_locale_dates_booleans_and_serialization() -> None:
    # Money helpers normalize supported input forms and reject empty/non-numeric values.
    assert to_pesos(12345) == 123.45
    assert to_cents(1500) == 150000
    assert to_cents(12.345) == 1234
    assert to_cents("1.500,50") == 150050
    assert to_cents("1500,50") == 150050
    assert to_positive_cents("-7.25") == 725
    with pytest.raises(ValueError, match="Empty amount"):
        to_cents("")
    with pytest.raises(ValueError):
        to_cents(True)

    # Dates accept common user/import formats, timezone-normalize, and reject invalid input.
    assert parse_date_input("2026-05-01").isoformat() == "2026-05-01T00:00:00+00:00"
    assert parse_date_input("01/05/26").isoformat() == "2026-05-01T00:00:00+00:00"
    assert parse_date_input("01-05-2026").isoformat() == "2026-05-01T00:00:00+00:00"
    assert parse_date_input("2026-05-01T03:00:00Z").tzinfo == UTC
    assert parse_date_input(datetime(2026, 5, 1)).tzinfo == UTC
    assert (
        parse_date_input(datetime(2026, 5, 1, 3, tzinfo=timezone.utc)).isoformat()
        == "2026-05-01T03:00:00+00:00"
    )
    with pytest.raises(ValueError, match="Invalid date value"):
        parse_date_input("")
    with pytest.raises(ValueError):
        parse_date_input("31/99/2026")

    assert parse_action_date("2026-05-01").tzinfo == UTC
    with pytest.raises(ValueError, match="Expected YYYY-MM-DD"):
        parse_action_date("01/05/2026")

    # Boolean, text, and ISO serialization helpers cover all accepted branches.
    assert parse_boolean(True) is True
    assert parse_boolean("yes") is True
    assert parse_boolean("0") is False
    with pytest.raises(ValueError, match="Invalid boolean value"):
        parse_boolean("perhaps")
    assert normalize_key("  Súper Mercado  ") == "super mercado"
    assert normalize_key(None) == ""
    assert normalize_text(None) == ""
    assert normalize_text("  Mixed Case  ") == "mixed case"
    assert isoformat_z(datetime(2026, 5, 1, 3)) == "2026-05-01T03:00:00Z"
    assert isoformat_z(datetime(2026, 5, 1, 3, tzinfo=timezone.utc)) == "2026-05-01T03:00:00Z"
    assert parse_date_input("2026-05-01T03:00:00").tzinfo == UTC


def test_remaining_import_document_settings_and_parser_edges(
    session: Session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Given importer helper inputs that are valid but uncommon.
    expenses = seed_category(session, name="Gastos")
    income = seed_category(session, name="Ingresos", kind=CategoryKind.INCOME)
    food = seed_subcategory(session, name="Comida", category=expenses)
    unique = seed_subcategory(session, name="Unica", category=expenses)
    seed_category(session, name="Otros")
    income_food = seed_subcategory(session, name="Comida", category=income)

    assert (
        import_service._fingerprint_payload(  # noqa: SLF001
            source_file="",
            external_id=None,
            source_row=None,
            date=datetime(2026, 5, 1, tzinfo=timezone.utc),
            amount_cents=100,
            business="Cafe",
            reason="Cafe",
            raw_description=None,
        )
        is None
    )
    assert import_service._bool_value({"reviewed": True}, "reviewed") is True  # noqa: SLF001
    assert import_service._bool_value({"reviewed": "yes"}, "reviewed") is True  # noqa: SLF001

    empty_structured = import_movements_structured(session, [], file_label="empty.json")
    assert empty_structured.to_dict()["inserted"] == 0

    # Structured batches report classification ambiguity and unknown references per row.
    structured = import_movements_structured(
        session,
        [
            {
                "_row_number": 1,
                "date": "2026-05-01",
                "amount": 100,
                "business": "Unique category fallback",
                "category": "Ingresos",
            },
            {
                "_row_number": 2,
                "date": "2026-05-02",
                "amount": 100,
                "business": "Unknown category id",
                "category_id": "missing-category",
            },
            {
                "_row_number": 3,
                "date": "2026-05-03",
                "amount": 100,
                "business": "Unknown category",
                "category": "Fantasma",
            },
            {
                "_row_number": 4,
                "date": "2026-05-04",
                "amount": 100,
                "business": "Unknown subcategory",
                "category_id": expenses.id,
                "subcategory": "Fantasma",
            },
            {
                "_row_number": 5,
                "date": "2026-05-05",
                "amount": 100,
                "business": "Ambiguous subcategory",
                "subcategory": "Comida",
            },
            {
                "_row_number": 6,
                "date": "2026-05-06",
                "amount": 100,
                "business": "Unique subcategory",
                "subcategory": "Unica",
            },
            {
                "_row_number": 7,
                "date": "2026-05-07",
                "amount": 100,
                "business": "Unknown subcategory without category",
                "subcategory": "Sin match",
            },
        ],
        file_label="edges.json",
    )
    assert structured.inserted == 3
    assert structured.failed == 4
    imported = session.get(Movement, structured.rows[0].movement_id)
    assert imported is not None
    assert imported.category_id == income.id
    unique_import = session.exec(
        select(Movement).where(Movement.business == "Unique subcategory")
    ).one()
    assert unique_import.subcategory_id == unique.id
    fallback_import = session.exec(
        select(Movement).where(Movement.business == "Unknown subcategory without category")
    ).one()
    assert fallback_import.category_id == expenses.id
    assert fallback_import.subcategory_id is None
    assert {error.row for error in structured.errors} == {2, 3, 4, 5}

    # CSV import keeps row-level failures and derives reason from business when needed.
    csv_outcome = import_movements_from_csv(
        session,
        _csv_text(
            [
                ["date", "amount", "business", "reason", "subcategory"],
                ["", "100", "Missing date", "", ""],
                ["2026-05-06", "200", "Business only", "", "Comida"],
            ]
        ),
        file_label="csv-edges.csv",
    )
    assert csv_outcome.inserted == 1
    assert csv_outcome.failed == 1
    business_only = session.exec(
        select(Movement).where(Movement.business == "Business only")
    ).one()
    assert business_only.reason == "Business only"
    assert business_only.subcategory_id in {food.id, income_food.id}

    # Document and XLSX helpers cover supported but less common branches.
    assert xlsx_service._stringify(date(2026, 5, 1)) == "2026-05-01"  # noqa: SLF001

    class EmptyWorkbook:
        active = None

    monkeypatch.setattr(
        xlsx_service.openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: EmptyWorkbook(),
    )
    assert xlsx_to_csv(b"anything") == ""

    class TextPage:
        def extract_text(self) -> str:
            return " page text "

    class TextPdf:
        pages = [TextPage()]

    monkeypatch.setattr(
        document_preprocessor,
        "PdfReader",
        lambda *_args, **_kwargs: TextPdf(),
    )
    pdf_doc = preprocess_document(
        filename="statement.pdf",
        content_type="application/pdf",
        data=b"%PDF",
    )
    assert pdf_doc.text == "--- Page 1 ---\npage text"
    assert pdf_doc.warning is None

    monkeypatch.setattr(settings_service, "get_primary_currency_code", lambda _session: "XXX")
    monkeypatch.setattr(
        settings_service,
        "currency_format_by_code",
        lambda code: {"code": "CLP"} if code == "CLP" else None,
    )
    assert settings_service.settings_payload(session)["primary_currency_code"] == "CLP"


def _csv_text(rows: list[list[str]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerows(rows)
    return buffer.getvalue()


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _blank_pdf_bytes() -> bytes:
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    buffer = io.BytesIO()
    writer.write(buffer)
    return buffer.getvalue()


def _force_subcategory_id(session: Session, movement_id: str, subcategory_id: str) -> None:
    session.commit()
    raw_connection = engine.raw_connection()
    try:
        cursor = raw_connection.cursor()
        cursor.execute("PRAGMA foreign_keys = OFF")
        cursor.execute(
            "UPDATE movement SET subcategory_id = ? WHERE id = ?",
            (subcategory_id, movement_id),
        )
        raw_connection.commit()
    finally:
        raw_connection.close()
    session.expire_all()


def _force_category_id(session: Session, category_id: str, business: str) -> None:
    session.commit()
    raw_connection = engine.raw_connection()
    try:
        cursor = raw_connection.cursor()
        cursor.execute("PRAGMA foreign_keys = OFF")
        cursor.execute(
            "UPDATE movement SET category_id = ? WHERE business = ?",
            (category_id, business),
        )
        raw_connection.commit()
    finally:
        raw_connection.close()
    session.expire_all()


def _seed_movement(
    session: Session,
    *,
    category_id: str,
    business: str,
    subcategory_id: str | None = None,
    reason: str | None = None,
    raw_description: str | None = None,
    amount_cents: int = 100000,
    reviewed: bool = False,
    source_file: str | None = "seed.csv",
    source_row: str | None = "1",
    import_hash: str | None = None,
) -> Movement:
    movement = Movement(
        date=datetime(2026, 5, 1, tzinfo=timezone.utc),
        accounting_date=datetime(2026, 5, 1, tzinfo=timezone.utc),
        amount_cents=amount_cents,
        business=business,
        reason=reason or business,
        source=MovementSource.MANUAL,
        raw_description=raw_description or reason or business,
        source_file=source_file,
        source_row=source_row,
        import_hash=import_hash,
        reviewed=reviewed,
        category_id=category_id,
        subcategory_id=subcategory_id,
    )
    session.add(movement)
    session.commit()
    session.refresh(movement)
    return movement
